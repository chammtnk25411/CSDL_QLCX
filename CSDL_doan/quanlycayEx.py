# quanlycayEx.py - Hoàn chỉnh với Sửa/Xóa
import sys
import os

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QMessageBox, QTableWidgetItem,
    QPushButton, QWidget, QHBoxLayout, QLabel, QDialog,
    QFileDialog, QTableWidget
)
from PyQt6.QtCore import Qt
from PyQt6.uic import loadUi

import pyodbc
import config

# Import dialog
from phieuthongtinEx import PhieuThongTinCayDialog


def get_db_connection():
    """Kết nối đến SQL Server"""
    try:
        drivers = [
            "ODBC Driver 18 for SQL Server",
            "ODBC Driver 17 for SQL Server",
            "SQL Server Native Client 11.0",
            "SQL Server"
        ]

        for driver in drivers:
            try:
                conn = pyodbc.connect(
                    f'DRIVER={{{driver}}};'
                    f'SERVER={config.DB_SERVER};'
                    f'DATABASE={config.DB_NAME};'
                    f'Trusted_Connection=yes;'
                    f'TrustServerCertificate=yes;'
                )
                return conn
            except:
                continue
        return None
    except:
        return None


class QuanLyCayWindow(QMainWindow):
    def __init__(self, username="", role=""):
        super().__init__()
        self.username = username
        self.role = role

        # ===== PHÂN QUYỀN =====
        self.is_guest = (role == "Khách tham quan")
        self.is_admin_or_staff = (role in ["Quản trị viên", "Nhân viên"])

        # Biến dữ liệu
        self.data = []
        self.filtered_data = []

        # Load UI
        ui_path = os.path.join(os.path.dirname(__file__), 'quanlycay.ui')
        try:
            loadUi(ui_path, self)
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể load UI: {str(e)}")
            sys.exit(1)

        # Cập nhật thông tin người dùng
        self.update_user_info()

        # Thiết lập bảng
        self.setup_table()

        # Kết nối sự kiện
        self.setup_connections()

        # Phân quyền
        self.setup_permissions()

        # Load dữ liệu
        self.load_data()

    def update_user_info(self):
        """Cập nhật thông tin người dùng"""
        if hasattr(self, "lbl_user_profile"):
            self.lbl_user_profile.setText(f"{self.username} ({self.role})")
        if hasattr(self, "sidebarUserLabel"):
            self.sidebarUserLabel.setText(f"👤 {self.username}")
        if hasattr(self, "sidebarRoleLabel"):
            self.sidebarRoleLabel.setText(self.role)
        if hasattr(self, "userLabel"):
            self.userLabel.setText(self.username)
        if hasattr(self, "roleLabel"):
            self.roleLabel.setText(self.role)

    def setup_table(self):
        """Thiết lập bảng - UI đã có 6 cột (có cột Thao tác)"""
        if not hasattr(self, "tableWidget"):
            return

        # Set headers
        headers = ["Mã cây", "Tên cây", "Loại thực vật", "Khu trưng bày", "Trạng thái", "Thao tác"]
        self.tableWidget.setHorizontalHeaderLabels(headers)

        # Set độ rộng cột
        self.tableWidget.setColumnWidth(0, 100)
        self.tableWidget.setColumnWidth(1, 200)
        self.tableWidget.setColumnWidth(2, 150)
        self.tableWidget.setColumnWidth(3, 150)
        self.tableWidget.setColumnWidth(4, 120)
        self.tableWidget.setColumnWidth(5, 130)

        self.tableWidget.verticalHeader().setVisible(False)
        self.tableWidget.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tableWidget.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.tableWidget.setRowCount(0)

    def setup_connections(self):
        """Kết nối sự kiện"""
        # Nút chức năng
        if hasattr(self, "btn_add"):
            self.btn_add.clicked.connect(self.open_add_dialog)
        if hasattr(self, "btn_export"):
            self.btn_export.clicked.connect(self.export_excel)

        # Tìm kiếm và lọc
        if hasattr(self, "txt_search"):
            self.txt_search.textChanged.connect(self.search_tree)
        if hasattr(self, "cb_khu"):
            self.cb_khu.currentIndexChanged.connect(self.filter_tree)
        if hasattr(self, "cb_trangthai"):
            self.cb_trangthai.currentIndexChanged.connect(self.filter_tree)

        # ===== SIDEBAR =====
        if hasattr(self, "pushButton"):
            self.pushButton.clicked.connect(self.open_trang_chu)
        if hasattr(self, "pushButton_2"):
            self.pushButton_2.clicked.connect(self.open_quan_ly_cay)
        if hasattr(self, "pushButton_3"):
            self.pushButton_3.clicked.connect(self.open_loai_thuc_vat)
        if hasattr(self, "pushButton_4"):
            self.pushButton_4.clicked.connect(self.open_ho_thuc_vat)
        if hasattr(self, "pushButton_5"):
            self.pushButton_5.clicked.connect(self.open_khu_trung_bay)
        if hasattr(self, "pushButton_6"):
            self.pushButton_6.clicked.connect(self.open_nhan_vien)
        if hasattr(self, "pushButton_7"):
            self.pushButton_7.clicked.connect(self.open_phieu_cham_soc)
        if hasattr(self, "pushButton_8"):
            self.pushButton_8.clicked.connect(self.open_phieu_khao_sat)
        if hasattr(self, "pushButton_9"):
            self.pushButton_9.clicked.connect(self.open_yeu_cau_bao_tri)
        if hasattr(self, "pushButton_10"):
            self.pushButton_10.clicked.connect(self.open_bao_cao_su_co)

    def setup_permissions(self):
        """Phân quyền"""
        if self.is_guest:
            if hasattr(self, "btn_add"):
                self.btn_add.setVisible(False)
                self.btn_add.setEnabled(False)
            if hasattr(self, "btn_export"):
                self.btn_export.setVisible(False)
                self.btn_export.setEnabled(False)
            if hasattr(self, "lbl_main_title"):
                self.lbl_main_title.setText("🔍 QUẢN LÝ CÂY - Chế độ xem")
            self.setWindowTitle("QUẢN LÝ CÂY - Chế độ xem")
        else:
            if hasattr(self, "btn_add"):
                self.btn_add.setVisible(True)
                self.btn_add.setEnabled(True)
            if hasattr(self, "btn_export"):
                self.btn_export.setVisible(True)
                self.btn_export.setEnabled(True)
            if hasattr(self, "lbl_main_title"):
                self.lbl_main_title.setText("QUẢN LÝ CÂY")
            self.setWindowTitle("QUẢN LÝ CÂY")

    def load_data(self):
        """Load dữ liệu từ database"""
        try:
            conn = get_db_connection()
            if not conn:
                QMessageBox.warning(self, "Lỗi", "Không thể kết nối database!")
                return

            cursor = conn.cursor()
            cursor.execute("""
                SELECT c.MACAY, c.TENCAY, l.TENTHUONGGOI, k.TENKHU, 
                       c.TRANGTHAIHOATDONG
                FROM CAY c
                LEFT JOIN LOAI_THUC_VAT l ON c.MALOAI = l.MALOAI
                LEFT JOIN KHU_TRUNG_BAY k ON c.MAKHU = k.MAKHU
                ORDER BY c.MACAY
            """)
            rows = cursor.fetchall()
            conn.close()

            self.data = []
            for row in rows:
                self.data.append({
                    'MACAY': row[0],
                    'TENCAY': row[1],
                    'TENLOAI': row[2] if row[2] else '',
                    'TENKHU': row[3] if row[3] else '',
                    'TRANGTHAI': row[4] if row[4] else ''
                })

            self.filtered_data = self.data.copy()
            self.display_data()
            print(f"✅ Đã load {len(self.data)} cây")

        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể load dữ liệu: {str(e)}")

    def display_data(self):
        """Hiển thị dữ liệu lên bảng"""
        if not hasattr(self, "tableWidget"):
            return

        self.tableWidget.setRowCount(0)

        for row_idx, item in enumerate(self.filtered_data):
            self.tableWidget.insertRow(row_idx)

            # 5 cột dữ liệu
            self.tableWidget.setItem(row_idx, 0, QTableWidgetItem(str(item.get('MACAY', ''))))
            self.tableWidget.setItem(row_idx, 1, QTableWidgetItem(str(item.get('TENCAY', ''))))
            self.tableWidget.setItem(row_idx, 2, QTableWidgetItem(str(item.get('TENLOAI', ''))))
            self.tableWidget.setItem(row_idx, 3, QTableWidgetItem(str(item.get('TENKHU', ''))))

            # Cột trạng thái
            trangthai = str(item.get('TRANGTHAI', ''))
            status_item = QTableWidgetItem(trangthai)
            status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.tableWidget.setItem(row_idx, 4, status_item)

            # ===== CỘT THAO TÁC =====
            if self.is_admin_or_staff:
                widget = QWidget()
                layout = QHBoxLayout(widget)
                layout.setContentsMargins(5, 2, 5, 2)
                layout.setSpacing(5)

                # Nút Sửa
                btn_edit = QPushButton("✏️ Sửa")
                btn_edit.setFixedSize(60, 28)
                btn_edit.setStyleSheet("""
                    QPushButton {
                        background-color: #2196F3;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        font-size: 11px;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background-color: #1976D2;
                    }
                """)
                btn_edit.clicked.connect(lambda checked, r=row_idx: self.edit_tree(r))

                # Nút Xóa
                btn_delete = QPushButton("🗑️ Xóa")
                btn_delete.setFixedSize(60, 28)
                btn_delete.setStyleSheet("""
                    QPushButton {
                        background-color: #f44336;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        font-size: 11px;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background-color: #d32f2f;
                    }
                """)
                btn_delete.clicked.connect(lambda checked, r=row_idx: self.delete_tree(r))

                layout.addWidget(btn_edit)
                layout.addWidget(btn_delete)
                layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

                self.tableWidget.setCellWidget(row_idx, 5, widget)
            else:
                label = QLabel("🔒 Chỉ xem")
                label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                label.setStyleSheet("""
                    QLabel {
                        color: #6c757d;
                        font-size: 12px;
                        font-weight: bold;
                        padding: 5px;
                    }
                """)
                self.tableWidget.setCellWidget(row_idx, 5, label)

    def search_tree(self):
        """Tìm kiếm cây"""
        if not hasattr(self, "txt_search"):
            return

        keyword = self.txt_search.text().strip().lower()

        if keyword == '':
            self.filtered_data = self.data.copy()
        else:
            self.filtered_data = [
                item for item in self.data
                if keyword in str(item.get('MACAY', '')).lower() or
                   keyword in str(item.get('TENCAY', '')).lower() or
                   keyword in str(item.get('TENLOAI', '')).lower() or
                   keyword in str(item.get('TENKHU', '')).lower()
            ]

        self.display_data()

    def filter_tree(self):
        """Lọc cây theo khu và trạng thái"""
        khu_filter = ""
        trangthai_filter = ""

        if hasattr(self, "cb_khu"):
            khu_filter = self.cb_khu.currentText()
        if hasattr(self, "cb_trangthai"):
            trangthai_filter = self.cb_trangthai.currentText()

        self.filtered_data = self.data.copy()

        if khu_filter and khu_filter != "Tất cả khu":
            self.filtered_data = [
                item for item in self.filtered_data
                if khu_filter == item.get('TENKHU', '')
            ]

        if trangthai_filter and trangthai_filter != "Tất cả trạng thái":
            self.filtered_data = [
                item for item in self.filtered_data
                if trangthai_filter == item.get('TRANGTHAI', '')
            ]

        self.display_data()

    def open_add_dialog(self):
        """Mở dialog thêm cây mới"""
        if self.is_guest:
            QMessageBox.warning(self, "⚠️ Cảnh báo", "Bạn không có quyền thêm cây mới!")
            return

        try:
            dialog = PhieuThongTinCayDialog(self, edit_mode=False)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                data = dialog.result_data
                if data:
                    try:
                        conn = get_db_connection()
                        if not conn:
                            QMessageBox.warning(self, "Lỗi", "Không thể kết nối database!")
                            return

                        cursor = conn.cursor()
                        cursor.execute("""
                            INSERT INTO CAY (MACAY, TENCAY, NGAYTRONG, CHIEUCAO, DUONGKINH,
                                             VITRI, TINHTRANGSINHTRUONG, TRANGTHAIHOATDONG,
                                             MALOAI, MAKHU)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            data['MACAY'],
                            data['TENCAY'],
                            data['NGAYTRONG'],
                            data['CHIEUCAO'],
                            data['DUONGKINH'],
                            data.get('VITRI', ''),
                            data['TINHTRANGSINHTRUONG'],
                            data['TRANGTHAIHOATDONG'],
                            data['MALOAI'],
                            data['MAKHU']
                        ))
                        conn.commit()
                        conn.close()

                        QMessageBox.information(self, "✅ Thành công", f"Đã thêm cây {data['TENCAY']}")
                        self.load_data()

                    except Exception as e:
                        QMessageBox.critical(self, "Lỗi", f"Không thể thêm cây: {str(e)}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi khi mở dialog: {str(e)}")

    def edit_tree(self, row):
        """Sửa thông tin cây"""
        print(f"🔧 edit_tree called with row: {row}")

        if self.is_guest:
            QMessageBox.warning(self, "⚠️ Cảnh báo", "Bạn không có quyền sửa thông tin cây!")
            return

        if row < 0 or row >= len(self.filtered_data):
            QMessageBox.warning(self, "Lỗi", f"Không tìm thấy cây tại vị trí {row}!")
            return

        item = self.filtered_data[row]
        print(f"📋 Dữ liệu cây: {item}")

        if not item.get('MACAY'):
            QMessageBox.warning(self, "Lỗi", "Mã cây không hợp lệ!")
            return

        try:
            dialog = PhieuThongTinCayDialog(self, edit_mode=True, data=item)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                data = dialog.result_data
                if data:
                    try:
                        conn = get_db_connection()
                        if not conn:
                            QMessageBox.warning(self, "Lỗi", "Không thể kết nối database!")
                            return

                        cursor = conn.cursor()
                        cursor.execute("""
                            UPDATE CAY 
                            SET TENCAY = ?, NGAYTRONG = ?, CHIEUCAO = ?, DUONGKINH = ?,
                                VITRI = ?, TINHTRANGSINHTRUONG = ?, TRANGTHAIHOATDONG = ?,
                                MALOAI = ?, MAKHU = ?
                            WHERE MACAY = ?
                        """, (
                            data['TENCAY'],
                            data['NGAYTRONG'],
                            data['CHIEUCAO'],
                            data['DUONGKINH'],
                            data.get('VITRI', ''),
                            data['TINHTRANGSINHTRUONG'],
                            data['TRANGTHAIHOATDONG'],
                            data['MALOAI'],
                            data['MAKHU'],
                            item['MACAY']
                        ))
                        conn.commit()
                        conn.close()

                        QMessageBox.information(self, "✅ Thành công", f"Đã cập nhật cây {data['TENCAY']}")
                        self.load_data()

                    except Exception as e:
                        QMessageBox.critical(self, "Lỗi", f"Không thể cập nhật cây: {str(e)}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi khi mở dialog: {str(e)}")

    def delete_tree(self, row):
        """Xóa cây"""
        print(f"🗑️ delete_tree called with row: {row}")

        if self.is_guest:
            QMessageBox.warning(self, "⚠️ Cảnh báo", "Bạn không có quyền xóa cây!")
            return

        if row < 0 or row >= len(self.filtered_data):
            QMessageBox.warning(self, "Lỗi", f"Không tìm thấy cây tại vị trí {row}!")
            return

        item = self.filtered_data[row]
        macay = item.get('MACAY', '')
        tencay = item.get('TENCAY', '')

        if not macay:
            QMessageBox.warning(self, "Lỗi", "Mã cây không hợp lệ!")
            return

        reply = QMessageBox.question(
            self,
            "⚠️ Xác nhận xóa",
            f"Bạn có chắc chắn muốn xóa cây '{tencay}' (Mã: {macay})?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                conn = get_db_connection()
                if not conn:
                    QMessageBox.warning(self, "Lỗi", "Không thể kết nối database!")
                    return

                cursor = conn.cursor()
                cursor.execute("DELETE FROM CAY WHERE MACAY = ?", (macay,))
                conn.commit()
                conn.close()

                QMessageBox.information(self, "✅ Thành công", f"Đã xóa cây '{tencay}'")
                self.load_data()

            except Exception as e:
                QMessageBox.critical(self, "Lỗi", f"Không thể xóa cây: {str(e)}")

    def export_excel(self):
        """Xuất dữ liệu ra Excel"""
        if self.is_guest:
            QMessageBox.warning(self, "⚠️ Cảnh báo", "Bạn không có quyền xuất dữ liệu!")
            return

        if not self.data:
            QMessageBox.warning(self, "Thông báo", "Không có dữ liệu để xuất!")
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment

            file_path, _ = QFileDialog.getSaveFileName(
                self, "Xuất Excel", "DanhSachCay.xlsx", "Excel Files (*.xlsx)"
            )
            if not file_path:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Danh sách cây"

            headers = ["Mã cây", "Tên cây", "Loại thực vật", "Khu trưng bày", "Trạng thái"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            for row_idx, item in enumerate(self.data, 2):
                ws.cell(row=row_idx, column=1, value=item.get('MACAY', ''))
                ws.cell(row=row_idx, column=2, value=item.get('TENCAY', ''))
                ws.cell(row=row_idx, column=3, value=item.get('TENLOAI', ''))
                ws.cell(row=row_idx, column=4, value=item.get('TENKHU', ''))
                ws.cell(row=row_idx, column=5, value=item.get('TRANGTHAI', ''))

            wb.save(file_path)
            QMessageBox.information(self, "✅ Thành công", f"Đã xuất file: {file_path}")

        except ImportError:
            QMessageBox.warning(self, "Thông báo", "Vui lòng cài đặt thư viện openpyxl để xuất Excel!")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể xuất Excel: {str(e)}")

    # ==================== CÁC HÀM MỞ TRANG ====================
    def open_trang_chu(self):
        try:
            from chinhEx import MainWindow as TrangChu
            self.window = TrangChu(self.username, self.role)
            self.window.show()
            self.close()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể mở trang chủ: {str(e)}")

    def open_quan_ly_cay(self):
        QMessageBox.information(self, "Thông báo", "Bạn đang ở trang Quản lý cây")

    def open_loai_thuc_vat(self):
        try:
            from LoaithucvatEx import MainWindow as LoaiThucVat
            self.window = LoaiThucVat(self.username, self.role)
            self.window.show()
            self.close()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể mở loài thực vật: {str(e)}")

    def open_ho_thuc_vat(self):
        try:
            from HothucvatEx import MainWindow as HoThucVat
            self.window = HoThucVat(self.username, self.role)
            self.window.show()
            self.close()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể mở họ thực vật: {str(e)}")

    def open_khu_trung_bay(self):
        try:
            from KhutrungbayEx import MainWindow as KhuTrungBay
            self.window = KhuTrungBay(self.username, self.role)
            self.window.show()
            self.close()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể mở khu trưng bày: {str(e)}")

    def open_nhan_vien(self):
        try:
            from NhanvienEx import MainWindow as NhanVien
            self.window = NhanVien(self.username, self.role)
            self.window.show()
            self.close()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể mở nhân viên: {str(e)}")

    def open_phieu_cham_soc(self):
        try:
            from phieu_cham_socEx import PhieuChamSocEx
            self.window = PhieuChamSocEx(self.username, self.role)
            self.window.show()
            self.close()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể mở phiếu chăm sóc: {str(e)}")

    def open_phieu_khao_sat(self):
        try:
            from phieu_khao_satEx import PhieuKhaoSatEx
            self.window = PhieuKhaoSatEx(self.username, self.role)
            self.window.show()
            self.close()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể mở phiếu khảo sát: {str(e)}")

    def open_yeu_cau_bao_tri(self):
        try:
            from yeu_cau_bao_triEx import YeuCauBaoTriEx
            self.window = YeuCauBaoTriEx(self.username, self.role)
            self.window.show()
            self.close()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể mở yêu cầu bảo trì: {str(e)}")

    def open_bao_cao_su_co(self):
        try:
            from bao_cao_su_coEx import MainWindow as BaoCaoSuCo
            self.window = BaoCaoSuCo(self.username, self.role)
            self.window.show()
            self.close()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể mở báo cáo sự cố: {str(e)}")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = QuanLyCayWindow(username="Admin", role="Quản trị viên")
    window.show()
    sys.exit(app.exec())